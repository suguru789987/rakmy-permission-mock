# -*- coding: utf-8 -*-
"""dataset/*.tsv から検証用データセットの xlsx を生成する。
TSV が唯一の元データ。シートを分けて投入しやすくし、先頭に使い方を置く。
"""
import csv,os,collections
from openpyxl import Workbook
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

ROOT=os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
D=os.path.join(ROOT,"dataset")
THIN=Side(style="thin",color="D6DEE5"); BOX=Border(THIN,THIN,THIN,THIN)
HDR=PatternFill("solid",fgColor="1F7A8C"); SUB=PatternFill("solid",fgColor="EDF3F7")
INK="1F4E5F"

def load(f): return list(csv.DictReader(open(os.path.join(D,f),encoding="utf-8"),delimiter="\t"))

def sheet(wb,name,rows,note="",widths=None):
    ws=wb.create_sheet(name)
    if note:
        ws.append([note]); ws["A1"].font=Font(size=10,color="5B6B7C")
        ws.append([])
    hdr=list(rows[0].keys()); ws.append(hdr); h=ws.max_row
    for r in rows: ws.append([r[k] for k in hdr])
    for c in ws[h]:
        c.font=Font(bold=True,size=10,color="FFFFFF"); c.fill=HDR
        c.alignment=Alignment(wrap_text=True,vertical="center",horizontal="center")
    for row in ws.iter_rows(min_row=h):
        for c in row:
            c.border=BOX
            if c.row>h: c.alignment=Alignment(wrap_text=True,vertical="top")
    for i,w in enumerate(widths or [22]*len(hdr),1): ws.column_dimensions[get_column_letter(i)].width=w
    ws.row_dimensions[h].height=30
    ws.freeze_panes=f"A{h+1}"
    return ws

wb=Workbook(); ws=wb.active; ws.title="使い方"
L=[("権限設定 検証用データセット","t"),
   ("検証プラン 20260804_権限設定_02_検証プラン.xlsx とセットで使います。すべて架空・ダミー（麺屋ミライ）です。本番データは使いません。","n"),
   ("",""),
   ("■ なぜデータが必要か","h"),
   ("データが空だと「塞げている」のか「データが無い」のか区別できません。権限設定の検証は「見えないこと」を確かめる場面が多く、空の画面はどちらとも解釈できます。","n"),
   ("",""),
   ("■ 金額を 1：2：4 にしている理由","h"),
   ("店舗ごとの金額を 1：2：4 にすると、合計値からどの店舗が含まれているか一意に逆算できます。","n"),
   ("100万＝渋谷のみ／200万＝新宿のみ／300万＝渋谷＋新宿（エリア長の正解）／400万＝池袋のみ／500万＝渋谷＋池袋／600万＝新宿＋池袋／700万＝全社（池袋が混入＝NG）","n"),
   ("合計を見るだけでどの店舗が混ざったか特定できます。比率を変えると逆算できなくなるので変えないでください。","n"),
   ("",""),
   ("■ 池袋三郎の給与だけ桁が違う理由","h"),
   ("エリア長の画面に 800,000 が現れたら、担当外店舗の従業員が見えている証拠になります。","n"),
   ("",""),
   ("■ 投入手順","h")]
for t,k in L:
    ws.append([t])
    r=ws.max_row
    if k=="t": ws.cell(r,1).font=Font(bold=True,size=14,color=INK)
    elif k=="h": ws.cell(r,1).font=Font(bold=True,size=11,color=INK); ws.cell(r,1).fill=SUB
    else: ws.cell(r,1).font=Font(size=10,color="333333")
    ws.cell(r,1).alignment=Alignment(wrap_text=True,vertical="center")
ws.append([]); ws.append(["順","やること","使うシート","対応する検証ID","済"]); hh=ws.max_row
STEP=[("0","権限テンプレート画面を確認（権限を作る前にしか出ない）","—","T-04",""),
      ("1","店舗3件を登録","01_店舗","S-02",""),
      ("2","売上・仕入・人件費・費用・予算を検証実施月の1日分として登録","03_売上／04_仕入／05_人件費／06_費用_予算","S-02",""),
      ("3","従業員4名を登録","02_従業員","S-02",""),
      ("4","検証ユーザー8名を登録","07_検証ユーザー","S-03",""),
      ("5","権限7件を作成","08_権限設定値","S-04",""),
      ("6","割当と担当店舗を設定","07_検証ユーザー","S-05",""),
      ("7","指標の初期値を突合（投入ではなく確認）","09_指標の初期値","T-19","")]
for s in STEP: ws.append(list(s))
for c in ws[hh]: c.font=Font(bold=True,size=10,color="FFFFFF"); c.fill=HDR; c.alignment=Alignment(horizontal="center",vertical="center")
for row in ws.iter_rows(min_row=hh,max_row=ws.max_row):
    for c in row: c.border=BOX; c.alignment=Alignment(wrap_text=True,vertical="top")
dv=DataValidation(type="list",formula1='"済,未,対象外"',allow_blank=True)
ws.add_data_validation(dv); dv.add(f"E{hh+1}:E{ws.max_row}")
ws.append([]); ws.append(["■ 注意"]); ws.cell(ws.max_row,1).font=Font(bold=True,size=11,color=INK); ws.cell(ws.max_row,1).fill=SUB
for t in ["日付は 2026-08-01 で作成しています。検証実施月に合わせて置換してください。",
          "実顧客のデータは使わないでください。検証で権限を絞る操作をするため、実データだと事故ります。",
          "投入できない場合、T-11・T-23 は「検証不能」として記録してください。「問題なし」と書いてはいけません。",
          "指標の既定は雛形で決まります。08 のベース列と 09 の期待値は必ず対応させてください。"]:
    ws.append([t]); ws.cell(ws.max_row,1).font=Font(size=10,color="333333")
for col,w in zip("ABCDE",[62,44,30,18,8]): ws.column_dimensions[col].width=w


# ── 目次シート（先頭に置く）
# 検証プランの「使う検証データ」列から、シートごとの検証IDを自動で集める
import re as _re
_TP=list(csv.DictReader(open(os.path.join(ROOT,"20260803_07_権限設定_検証プラン_PdM引継.tsv"),encoding="utf-8"),delimiter="\t"))
def _tids(sheet):
    key=sheet.split("_")[0]; out=[]
    for _x in _TP:
        v=_re.sub(r"US-\d+","",_x.get("使う検証データ",""))
        if key in _re.findall(r"(?<![-\d])(\d{2})(?=[_（\s／]|$)",v): out.append(_x["検証ID"])
    return " ".join(out) or "—"
IDX=[("01_店舗","店舗3件（渋谷・新宿・池袋）","投入",_tids("01"),"池袋は担当外＝見えてはいけない店舗として使う"),
 ("02_従業員","従業員4名と月給","投入",_tids("02"),"池袋三郎だけ給与80万。エリア長の画面に出たら担当外が見えている証拠"),
 ("03_売上","売上・客数・客単価","投入",_tids("03"),"1：2：4。合計からどの店舗が混ざったか逆算できる"),
 ("04_仕入","仕入と原価率","投入",_tids("04"),"原価率を全店10%に統一。項目の有無だけを見るため"),
 ("05_人件費","店舗別の人件費","投入",_tids("05"),"1：2：4"),
 ("06_費用_予算","地代家賃と売上予算","投入",_tids("06"),"集計・予実対比の画面を空にしないため。予算は売上の1.1倍"),
 ("07_検証ユーザー","検証ユーザーと割当権限・担当店舗","投入",_tids("07"),"US-07 は付け替え検証のため「エリア長 → 経理（会社）」の遷移表記"),
 ("08_権限設定値","権限ごとの画面設定値","投入",_tids("08"),"そのまま入力できる。ベース列は指標の既定を決める"),
 ("09_指標の初期値","権限ごとの指標ON件数","突合",_tids("09"),"投入データではない。実測値と突き合わせる期待値"),
]
wi=wb.create_sheet("目次",0)
wi.append(["権限設定 検証用データセット　目次"])
wi["A1"].font=Font(bold=True,size=14,color=INK)
wi.append(["投入＝ステージングに登録するデータ／突合＝登録せず実測値と突き合わせる期待値"])
wi["A2"].font=Font(size=10,color="5B6B7C")
wi.append([])
wi.append(["シート","内容","件数","種別","使う検証ID","このデータの役割"]); hi=wi.max_row
for name,desc,kind,tids,role in IDX:
    n=len(load(name+".tsv"))
    extra=""
    if name=="08_権限設定値":
        import collections as _c
        g=_c.OrderedDict()
        for x in load(name+".tsv"): g.setdefault(x["権限名"],0)
        extra=f"（権限{len(g)}件）"
    wi.append([name,desc,f"{n}行{extra}",kind,tids,role])
for c in wi[hi]:
    c.font=Font(bold=True,size=10,color="FFFFFF"); c.fill=HDR
    c.alignment=Alignment(wrap_text=True,vertical="center",horizontal="center")
for row in wi.iter_rows(min_row=hi,max_row=wi.max_row):
    for c in row:
        c.border=BOX
        if c.row>hi: c.alignment=Alignment(wrap_text=True,vertical="top")
for col,w in zip("ABCDEF",[18,32,16,10,24,50]): wi.column_dimensions[col].width=w
wi.row_dimensions[hi].height=30
for r in range(hi+1,wi.max_row+1): wi.row_dimensions[r].height=34
wi.freeze_panes=f"A{hi+1}"
wi.append([])
wi.append(["■ 読む順番"]); wi.cell(wi.max_row,1).font=Font(bold=True,size=11,color=INK); wi.cell(wi.max_row,1).fill=SUB
for t in ["1. 「使い方」で なぜこのデータが必要か・1：2：4 の理由 を読む",
          "2. 「投入手順」のとおり 01 → 03〜06 → 02 → 07 → 08 → 07 の順に登録する",
          "3. 検証を実施し、09 の期待値と実測値を突き合わせる（T-19）"]:
    wi.append([t]); wi.cell(wi.max_row,1).font=Font(size=10,color="333333")

NOTE={"01_店舗.tsv":"池袋店は「担当外＝見えてはいけない店舗」として使います。",
 "02_従業員.tsv":"池袋三郎の給与だけ桁を上げています。エリア長の画面に 800,000 が出たら担当外が見えている証拠です。",
 "03_売上.tsv":"1：2：4。合計からどの店舗が含まれるか逆算できます。比率を変えないでください。",
 "04_仕入.tsv":"原価率を全店10%で揃えています。T-12 で「原価率」という項目の有無だけを見たいためです。",
 "05_人件費.tsv":"1：2：4。","06_費用_予算.tsv":"予算は集計・予実対比の画面を空にしないために入れます。売上の1.1倍です。",
 "07_検証ユーザー.tsv":"US-07 は付け替え検証（T-15）のため「エリア長 → 経理（会社）」の遷移で記載しています。",
 "08_権限設定値.tsv":"S-04 でそのまま入力できます。ベース列は指標の既定を決めるため 09 と必ず対応させてください。",
 "09_指標の初期値.tsv":"投入データではなく突合用の期待値です。T-19 で使います。"}
W={"07_検証ユーザー.tsv":[10,18,32,26,26,22],"08_権限設定値.tsv":[10,24,8,34,26,10,12],
   "09_指標の初期値.tsv":[26,8,12,12,40,18],"02_従業員.tsv":[18,22,12,12,12,20],
   "06_費用_予算.tsv":[12,22,14,14,10]}
for f in sorted(os.listdir(D)):
    if not f.endswith(".tsv"): continue
    sheet(wb,f[:-4],load(f),NOTE.get(f,""),W.get(f))

p=os.path.join(ROOT,"20260806_権限設定_03_検証用データセット.xlsx")
wb.save(p)
print(f"  {os.path.basename(p)}  シート{len(wb.sheetnames)}枚: {wb.sheetnames}")
