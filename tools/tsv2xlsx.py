# -*- coding: utf-8 -*-
"""受入条件表・検証プランの TSV から xlsx を再生成する。
TSV が唯一の元データ。書式（集計ブロック・条件付き書式・入力規則）はここで付ける。
"""
import csv,sys,os
from openpyxl import Workbook
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

ROOT=os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
THIN=Side(style="thin",color="D6DEE5")
BOX=Border(THIN,THIN,THIN,THIN)
HDR=PatternFill("solid",fgColor="1F7A8C")
SUB2=PatternFill("solid",fgColor="EDF3F7")

def style(ws,hrow,ncol,widths,rowh=64):
    for c in ws[hrow]:
        c.font=Font(bold=True,size=10,color="FFFFFF"); c.fill=HDR
        c.alignment=Alignment(wrap_text=True,vertical="center",horizontal="center")
    for row in ws.iter_rows(min_row=hrow):
        for c in row:
            c.border=BOX
            if c.row>hrow: c.alignment=Alignment(wrap_text=True,vertical="top")
    for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
    for r in range(hrow+1,ws.max_row+1): ws.row_dimensions[r].height=rowh
    ws.row_dimensions[hrow].height=34

def build_ac():
    rows=list(csv.DictReader(open(f"{ROOT}/20260803_06_権限設定_受入条件表_PdM引継.tsv",encoding="utf-8"),delimiter="\t"))
    hdr=list(rows[0].keys()); wb=Workbook(); ws=wb.active; ws.title="受入条件表"
    ws["A1"]="権限設定 受入条件表 ─ どこまで実装できれば合格か"
    ws["A2"]="「合否」に OK / NG を入れると下の到達状況が自動で集計されます。判定は上のマイルストーン順に進めてください。"
    ws["A1"].font=Font(bold=True,size=13,color="1F4E5F"); ws["A2"].font=Font(size=10,color="5B6B7C")
    ws.append([])
    ms=[]
    for r in rows:
        if r["マイルストーン"] not in [m[0] for m in ms]: ms.append((r["マイルストーン"],r["マイルストーンの完了条件"]))
    ws.append(["マイルストーン","完了条件","","","条件数","OK","NG","未判定","到達率"])
    hrow_sum=ws.max_row
    first=hrow_sum+len(ms)+4   # データ開始行（集計ヘッダ＋マイルストーン行＋空行2＋見出し行の次）
    for m,cond in ms:
        r=ws.max_row+1
        ws.append([m,cond,"","",
                   f'=COUNTIF($A${first}:$A${first+len(rows)-1},A{r})',
                   f'=COUNTIFS($A${first}:$A${first+len(rows)-1},A{r},$N${first}:$N${first+len(rows)-1},"OK")',
                   f'=COUNTIFS($A${first}:$A${first+len(rows)-1},A{r},$N${first}:$N${first+len(rows)-1},"NG")',
                   f"=E{r}-F{r}-G{r}", f'=IF(E{r}=0,"",F{r}/E{r})'])
    for c in ws[hrow_sum]: c.font=Font(bold=True,size=10,color="FFFFFF"); c.fill=HDR
    for r in range(hrow_sum,ws.max_row+1):
        for c in ws[r]: c.border=BOX; c.alignment=Alignment(wrap_text=True,vertical="center")
        ws.cell(row=r,column=9).number_format="0%"
    ws.append([]); ws.append([])
    ws.append(hdr); hrow=ws.max_row
    for r in rows: ws.append([r[k] for k in hdr])
    ni=hdr.index("合否")+1
    dv=DataValidation(type="list",formula1='"OK,NG,対象外"',allow_blank=True)
    ws.add_data_validation(dv); dv.add(f"{get_column_letter(ni)}{hrow+1}:{get_column_letter(ni)}{ws.max_row}")
    rng=f"{get_column_letter(ni)}{hrow+1}:{get_column_letter(ni)}{ws.max_row}"
    ws.conditional_formatting.add(rng,CellIsRule(operator="equal",formula=['"OK"'],fill=PatternFill("solid",fgColor="D8F0E0")))
    ws.conditional_formatting.add(rng,CellIsRule(operator="equal",formula=['"NG"'],fill=PatternFill("solid",fgColor="FBDCDC")))
    W=[16,30,16,9,10,34,40,36,22,22,16,26,24,9,16,12,12]
    style(ws,hrow,len(hdr),W,rowh=76)
    ws.freeze_panes=f"D{hrow+1}"
    p=f"{ROOT}/20260804_権限設定_01_受入条件表.xlsx"; wb.save(p); return p,len(rows)

def build_tp():
    rows=list(csv.DictReader(open(f"{ROOT}/20260803_07_権限設定_検証プラン_PdM引継.tsv",encoding="utf-8"),delimiter="\t"))
    hdr=list(rows[0].keys()); wb=Workbook(); ws=wb.active; ws.title="検証プラン"
    ws["A1"]="確かめる手順。上から順に実施する。S-01〜S-06 は準備、T-01〜 が検証。\n「画面」を開いて「操作」を行い、「期待挙動」と「起きてはいけないこと」を見て、「合否ライン」に照らして「判定」に OK / NG を入れる。\n右端の「受入条件ID／遷移フローID／ヘルプページの該当箇所」は、NG のときにどこへ跳ね返るかを示す。"
    ws["A1"].font=Font(bold=True,size=11,color="1F4E5F")
    ws.append(hdr); hrow=ws.max_row
    for r in rows: ws.append([r[k] for k in hdr])
    ji=hdr.index("判定")+1
    dv=DataValidation(type="list",formula1='"OK,NG,検証不能"',allow_blank=True)
    ws.add_data_validation(dv); dv.add(f"{get_column_letter(ji)}{hrow+1}:{get_column_letter(ji)}{ws.max_row}")
    rng=f"{get_column_letter(ji)}{hrow+1}:{get_column_letter(ji)}{ws.max_row}"
    ws.conditional_formatting.add(rng,CellIsRule(operator="equal",formula=['"OK"'],fill=PatternFill("solid",fgColor="D8F0E0")))
    ws.conditional_formatting.add(rng,CellIsRule(operator="equal",formula=['"NG"'],fill=PatternFill("solid",fgColor="FBDCDC")))
    style(ws,hrow,len(hdr),[8,9,22,26,24,38,36,34,40,36,34,10,16,12,32,16,9],rowh=120)
    ws.freeze_panes=f"C{hrow+1}"

    # ── 列仕様シート
    ws2=wb.create_sheet("列仕様",0)
    ws2["A1"]="検証プラン 列仕様"
    ws2["A1"].font=Font(bold=True,size=14,color="1F4E5F")
    ws2["A2"]="計算系と操作系を統合した1枚です。計算系だけに使う列（入力値・計算根拠）は、操作系では「—」になります。"
    ws2["A2"].font=Font(size=10,color="5B6B7C")
    ws2.append([])
    ws2.append(["#","列名","何を書くか","記入者","例"]); h2=ws2.max_row
    SPECROWS=[
 ("1","段階","準備／設定／権限が効くか／割当／一覧・遷移／指標／移植 のどれか。上から順に実施する","作成時","権限が効くか"),
 ("2","検証ID","S＝準備、T＝検証。実施順に並べている","作成時","T-11"),
 ("3","検証内容","何を確かめるケースか。ひと言で","作成時","US-03 エリア長（店長ベースのカスタマイズ例）"),
 ("4","画面","どの画面を開くか。ログインするユーザーも書く","作成時","US-03でログイン → 売上分析・従業員管理"),
 ("5","前提・入力値","先に済んでいる必要のある準備ID。計算系はここに入力値を書く","作成時","S-05／入力値：担当店舗＝渋谷店・新宿店"),
 ("6","操作","実際に行う手順。クリックする場所まで書く","作成時","US-03でログインし、許可画面を順に開く→次に不許可画面をメニュー確認し直URLでも開く"),
 ("7","期待挙動","何が起きればよいか（できること）","作成時","担当2店舗の合計＝3,000,000 になる"),
 ("8","起きてはいけないこと","何が出たらNGか（見えてはいけないもの）。権限の検証は「見えないこと」を確かめる場面が多いため独立させている","作成時","池袋（400万）が混ざらない。池袋三郎の給与80万が出ない"),
 ("9","合否ライン","OK/NGを分ける数値の線。件数・割合・金額で書く","作成時","集計の合計＝3,000,000。池袋の混入＝0件（7,000,000ならNG）"),
 ("10","計算根拠","その数値になる理由。計算系のみ。操作系は「—」","作成時","渋谷100万＋新宿200万＝300万。池袋400万が入ると700万になる"),
 ("11","確認ポイント","なぜこれを見るのか、どこで間違えやすいか","作成時","合計が7,000,000なら全社が見えている＝NG。P3を前倒しする根拠"),
 ("12","スクショファイル名","証跡の保存名。検証IDと合わせる","実施時","T-11"),
 ("13","受入条件ID","この検証が担保する受入条件。NGならリリース判断に跳ね返る","作成時","AC-10 AC-11 AC-23"),
 ("14","遷移フローID","対応する遷移フロー（20260803_09）。NGなら遷移の設計に跳ね返る","作成時","F-14"),
 ("15","ヘルプページの該当箇所","顧客に何と説明しているか。NGならヘルプの記述も直す必要がある","作成時","ケース別のやり方（複数の店舗をまとめて見せたい）／担当店舗を設定する（図2）"),
 ("16","実際の値","実施して見えた値・挙動をそのまま書く。判断は書かない","実施時","合計 3,000,000。池袋は切替に出ず"),
 ("17","判定","OK／NG／検証不能。データを投入できない場合は「検証不能」。「問題なし」と書かない","実施時","OK"),
    ]
    for r in SPECROWS: ws2.append(list(r))
    for c in ws2[h2]:
        c.font=Font(bold=True,size=10,color="FFFFFF"); c.fill=HDR
        c.alignment=Alignment(wrap_text=True,vertical="center",horizontal="center")
    for row in ws2.iter_rows(min_row=h2,max_row=ws2.max_row):
        for c in row:
            c.border=BOX
            if c.row>h2: c.alignment=Alignment(wrap_text=True,vertical="top")
    for col,w in zip("ABCDE",[5,22,54,10,52]): ws2.column_dimensions[col].width=w
    for r in range(h2+1,ws2.max_row+1): ws2.row_dimensions[r].height=42
    ws2.freeze_panes=f"A{h2+1}"
    ws2.append([]); ws2.append(["■ 使い方"])
    ws2.cell(ws2.max_row,1).font=Font(bold=True,size=11,color="1F4E5F"); ws2.cell(ws2.max_row,1).fill=SUB2
    for t in ["1. 「画面」を開き、「前提・入力値」が済んでいることを確かめる",
              "2. 「操作」のとおりに操作する",
              "3. 「期待挙動」と「起きてはいけないこと」の両方を見る",
              "4. 「合否ライン」に照らして「実際の値」と「判定」を書く",
              "5. NG のときは「受入条件ID」「遷移フローID」「ヘルプページの該当箇所」を見て、どこに跳ね返るかを確認する"]:
        ws2.append([t]); ws2.cell(ws2.max_row,1).font=Font(size=10,color="333333")

    p=f"{ROOT}/20260804_権限設定_02_検証プラン.xlsx"; wb.save(p); return p,len(rows)

if __name__=="__main__":
    for p,n in (build_ac(),build_tp()): print(f"  {os.path.basename(p)}  {n}件")
