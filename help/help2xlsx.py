# -*- coding: utf-8 -*-
import re,csv,os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
lines=open("help/draft-framer.md",encoding="utf-8").read().split("\n")
rows=[]  # [階層, 内容, 補足]
def add(lv,c,n=""): rows.append([lv,c,n])
STOP=r"^(#|【|\*\*|⚠️|\||---)"
def clean(t): return re.sub(r"\*\*(.+?)\*\*",r"\1",t).replace("**","")
i=0; guard=0
while i<len(lines):
    guard+=1
    if guard>5000: break
    l=lines[i].rstrip(); start=i
    if not l.strip(): i+=1; continue
    if l.strip()=="---": i+=1; continue
    if l.startswith("# "): add("H1",clean(l[2:])); i+=1; continue
    if l.startswith("## "): add("H2",clean(l[3:])); i+=1; continue
    if l.startswith("### "): add("H3",clean(l[4:])); i+=1; continue
    m=re.match(r"^【画像([①-⑩])：(.+)】$",l)
    if m: add("画像",f"【画像{m.group(1)}】",clean(m.group(2))); i+=1; continue
    if l.startswith("> "):
        blk=[]
        while i<len(lines) and lines[i].startswith(">"):
            blk.append(re.sub(r"^>\s?","",lines[i]).strip()); i+=1
        head=clean(blk[0]) if blk else ""
        add("注記（引用）","> "+head,clean(" ".join(blk[1:])))
        continue
    m=re.match(r"^⚠️ \*\*(.+?)\*\*$",l)
    if m:
        h=clean(m.group(1)); i+=1; b=[]
        while i<len(lines) and lines[i].strip() and not re.match(STOP,lines[i]): b.append(lines[i].strip()); i+=1
        add("注記（引用）","> "+h,clean(" ".join(b))); continue
    if l.strip().startswith("**目次") and l.strip().endswith("**"):
        add("目次",clean(l.strip())); i+=1; continue
    m=re.match(r"^\*\*Q\. (.+?)\*\*$",l)
    if m:
        a=lines[i+1].strip() if i+1<len(lines) else ""
        add("FAQ","Q. "+clean(m.group(1)),clean(re.sub(r"^A\. ","",a))); i+=2; continue
    m=re.match(r"^\*\*([^*]+?)\*\*$",l)
    if m:
        h=clean(m.group(1)); i+=1; b=[]
        while i<len(lines) and lines[i].strip() and not re.match(STOP,lines[i]): b.append(lines[i].strip()); i+=1
        add("小見出し",h,clean(" ".join(b))); continue
    if l.startswith("|"):
        tbl=[]
        while i<len(lines) and lines[i].startswith("|"):
            cs=[clean(c.strip()) for c in lines[i].strip("|").split("|")]
            if not all(re.fullmatch(r":?-{2,}:?",c) for c in cs): tbl.append(cs)
            i+=1
        if tbl:
            add("表 見出し"," ｜ ".join(tbl[0]))
            for r in tbl[1:]: add("表 行"," ｜ ".join(r))
        if i==start: i+=1
        continue
    if re.match(r"^\d+\. ",l) or l.startswith("- "):
        while i<len(lines) and (re.match(r"^\d+\. ",lines[i]) or lines[i].startswith("- ")):
            t=lines[i]; num=bool(re.match(r"^\d",t))
            add("手順" if num else "箇条書き",clean(re.sub(r"^- ","・",t).strip())); i+=1
        continue
    b=[l.strip()]; i+=1
    while i<len(lines) and lines[i].strip() and not re.match(STOP+r"|^- |^\d+\. ",lines[i]): b.append(lines[i].strip()); i+=1
    add("本文",clean(" ".join(b)))
HDR=PatternFill("solid",fgColor="1F4E5F"); HF=Font(color="FFFFFF",bold=True,size=10)
THIN=Side(style="thin",color="D0D7DE"); BD=Border(left=THIN,right=THIN,top=THIN,bottom=THIN)
WRAP=Alignment(wrap_text=True,vertical="top")
TONE={"目次":"F5F7F9","H1":"1F4E5F","H2":"D6E7EF","H3":"E8F1F5","注記（引用）":"FFF6E5","小見出し":"F0F4F7","画像":"EEF6FB","FAQ":"F5F5F5","表 見出し":"EDF3F7"}
wb=Workbook(); ws=wb.active; ws.title="ヘルプページ"
ws.append(["権限設定 ヘルプページ原稿　help.rakmy.jp / management / 権限設定　※ページ掲載時の見た目の順どおりに並べています。画像10枚は未取得。"])
ws.cell(row=1,column=1).font=Font(size=9,color="6B7280",italic=True)
ws.append(["階層","内容","補足"])
for r in rows: ws.append(r)
for c in range(1,4):
    x=ws.cell(row=2,column=c); x.fill=HDR; x.font=HF
    x.alignment=Alignment(wrap_text=True,vertical="center",horizontal="center")
for rr in range(3,ws.max_row+1):
    k=ws.cell(row=rr,column=1).value
    for c in range(1,4):
        cell=ws.cell(row=rr,column=c); cell.border=BD; cell.alignment=WRAP
        t=TONE.get(k)
        if t: cell.fill=PatternFill("solid",fgColor=t)
    if k=="H1":
        for c in range(1,4): ws.cell(row=rr,column=c).font=Font(bold=True,size=13,color="FFFFFF")
    elif k=="H2": ws.cell(row=rr,column=2).font=Font(bold=True,size=12,color="1F4E5F")
    elif k=="H3": ws.cell(row=rr,column=2).font=Font(bold=True,size=11,color="1F4E5F")
    elif k=="表 見出し": ws.cell(row=rr,column=2).font=Font(bold=True,size=10)
    elif k=="注記（引用）": ws.cell(row=rr,column=2).font=Font(bold=True,size=10,color="8A6D0B")
    elif k=="小見出し": ws.cell(row=rr,column=2).font=Font(bold=True,size=10,color="1F4E5F")
for c in range(1,4): ws.cell(row=2,column=c).border=BD
ws.freeze_panes="A3"; ws.auto_filter.ref=f"A2:C{ws.max_row}"
for i,w in enumerate([14,78,60],1): ws.column_dimensions[get_column_letter(i)].width=w
OUT=os.path.expanduser("~/Desktop/権限設定_20260804/ヘルプページ")
wb.save(f"{OUT}/20260805_権限設定_ヘルプページ.xlsx"); wb.save("20260805_権限設定_ヘルプページ.xlsx")
with open(f"{OUT}/20260805_権限設定_ヘルプページ.csv","w",encoding="utf-8-sig",newline="") as f:
    csv.writer(f).writerows([["階層","内容","補足"]]+rows)
from collections import Counter
print(f"作成: {len(rows)}行")
print("  階層:",dict(Counter(r[0] for r in rows)))
